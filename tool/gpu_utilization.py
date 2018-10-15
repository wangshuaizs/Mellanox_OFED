# -*- coding: utf-8 -*-  
from __future__ import print_function

import numpy as np  
import argparse
import matplotlib.pyplot as plt 
import socket
import linecache
from openpyxl import Workbook
from openpyxl import load_workbook
import time

def main():
    parser = argparse.ArgumentParser(description="progrom description")
    parser.add_argument('-f', '--file', type=str, default="results-file.xlsx")
    parser.add_argument('-x', '--xlsxfile', type=str, default="gpu_utilization.xlsx")
    parser.add_argument('-i', '--interval', type=int, default=1)
    parser.add_argument('-l', '--min', type=float, default=0)
    parser.add_argument('-r', '--max', type=float)
    parser.add_argument('-hi', '--high', type=float)
    parser.add_argument('-t', '--time', type=int)
    args = parser.parse_args()
    rawfile = args.file
    interval = args.interval
    xlsxfile = args.xlsxfile
    axis_min = args.min
    axis_max = args.max
    axis_high = args.high
    comm_time_0 = args.time // 1000000
    
    current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    current_time_list = current_time.split(":")
    time_00_str = current_time_list[0] + ":00:00"
    time_00 = int(time.mktime(time.strptime(time_00_str,"%Y-%m-%d %H:%M:%S")))
    
    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet["A1"].value = "time/ms"
    new_sheet["B1"].value = "gpu utilization/%"
    new_sheet_cur = 2

    wb = load_workbook(rawfile)
    sheet = wb['Sheet1']
    max_row = sheet.max_row
    has_skipped_the_first = 0
    value = sheet['A2'].value
    start_row = 2
    relative_time = 0
    print(max_row)
    for i in xrange(2, max_row) :
        tmp = str(sheet['A'+str(i)].value)
        if tmp.find(".") != -1 :
            timestamp = int(time.mktime(time.strptime(tmp,"%Y-%m-%d %H:%M:%S.%f"))) * 1000 + int(tmp.split(".")[1][0:3])
        else :
            timestamp = int(time.mktime(time.strptime(tmp,"%Y-%m-%d %H:%M:%S"))) * 1000
        relative_time = timestamp - comm_time_0
        new_sheet['A'+str(new_sheet_cur)].value = relative_time
        new_sheet['B'+str(new_sheet_cur)].value = sheet['C'+str(i)].value * 100
        new_sheet_cur = new_sheet_cur + 1
    new_wb.save(xlsxfile)

if __name__ == '__main__':
    main()